from openpyxl import Workbook


def process_38704_results(result_data: str):
    rows = result_data.split('\n')
    protocols = []
    protocol = {'KeyExchangeMethods': [], 'Name': ''}
    method = {}
    for row in rows:
        if row.split('\t')[0] == 'CIPHER' or row.split('\t')[0] == 'NAME':
            # This is the column header, skip it
            continue
        elif len([c for c in row.split('\t') if c != ' ']) == 1:
            # This is a protocol header row, so commit and reset the protocol
            if len(protocol['KeyExchangeMethods']) > 0:
                protocols.append(protocol)
                protocol = {'Name': row.split('\t')[0], 'KeyExchangeMethods': []}
        else:
            columns = row.split('\t')
            # If the number of columns is 7, the cipher name is included.  If the number of columns is 6,
            # the cipher name is not included
            if len(columns) == 6:
                method = {
                    'Name': columns[0],
                    'Group': columns[1],
                    'KeySize': columns[2],
                    'ForwardSecret': columns[3],
                    'ClassicalStrength': columns[4],
                    'QuantumStrength': columns[5]
                }
            elif len(columns) == 7:
                method = {
                    'Cipher': columns[0],
                    'Name': columns[1],
                    'Group': columns[2],
                    'KeySize': columns[3],
                    'ForwardSecret': columns[4],
                    'ClassicalStrength': columns[5],
                    'QuantumStrength': columns[6]
                }
            else:
                # If we haven't been able to handle the row, discard it and move on to the next
                continue
            protocol['KeyExchangeMethods'].append(method)
    return protocols

def process_38116_results(result_data: str):
    rows = result_data.split('\n')
    protocols = []
    protocol = {}
    for row in rows:
        # Each row (except the column header row) is one of 3 things:
        #   1.  A protocol header which shows the protocol name and its status (enabled or disabled)
        #       e.g. (TLSv1 PROTOCOL IS ENABLED)
        #   2.  The protocol Compression Method (only available if the protocol is enabled)
        #       e.g. (TLSv1\tCOMPRESSION METHOD\tNone)
        #   3.  A ciphersuite entry available in the protocol (only available if the protocol is enabled)
        #       Data is presented as 'CIPHER\tKEY-EXCHANGE\tAUTHENTICATION\tMAC\tENCRYPTION(KEY-STRENGTH)\tGRADE
        #       e.g. RC4-MD5\tRSA\tRSA\tMD5\tRC4(128)\tMEDIUM

        if row.split('\t')[0] == 'CIPHER':
            # This is the column header row, so we can skip it
            continue
        elif row.find('PROTOCOL IS') > -1:
            # Type 1 row, meaning this is a new protocol.  If we are currently processing a protocol, we now need
            # to commit that one and start a new one.
            if len(protocol) > 0:
                protocols.append(protocol)
                protocol = {}
            protocol['Name'] = row.split(' ')[0]
            protocol['Status'] = row.split(' ')[3].strip('\t')
            protocol['CompressionMethod'] = ''
            protocol['Ciphersuite'] = []
        elif row.find('COMPRESSION METHOD') > -1:
            # Type 2 row, add the compression method to the protocol dict
            protocol['CompressionMethod'] = row.split('\t')[2]
        else:
            cipher_data = row.split('\t')
            if cipher_data[4].find('(') > -1:
                encryption = cipher_data[4].split('(')[0]
                keystrength = cipher_data[4].split('(')[1].strip(')')
            else:
                encryption = cipher_data[4]
                keystrength = ''

            cipher = {
                'Name': cipher_data[0],
                'KeyExchange': cipher_data[1],
                'Authentication': cipher_data[2],
                'MAC': cipher_data[3],
                'Encryption': encryption,
                'KeyStrength': keystrength,
                'Grade': cipher_data[5]
            }
            protocol['Ciphersuite'].append(cipher)
    return protocols


def process_38116_detection(detection: dict):
    detection_data = {
        'TLS': {
            '38116': {
                'Port': detection['PORT'],
                'LastDetected': detection['LAST_FOUND_DATETIME'],
                'Protocols': process_38116_results(detection['RESULTS'])
            }
        }
    }
    return detection_data


def process_38704_detection(detection: dict):
    detection_data = {
        'TLS': {
            '38704': {
                'Port': detection['PORT'],
                'LastDetected': detection['LAST_FOUND_DATETIME'],
                'Protocols': process_38704_results(detection['RESULTS'])
            }
        }
    }
    return detection_data


def process_tls_detections(data: list):
    # The actual detection data is buried inside the returned XML (now JSON)
    # ['HOST_LIST_VM_DETECTION_OUTPUT']['RESPONSE']['HOST_LIST']['HOST'] is an array of hosts
    #
    # ['DETECTION_LIST'] is a single element, not a list, containing a single ['DETECTION'] element
    # ['DETECTION'] is either a list of  detection elements for the host or a single element containing one detection
    # ['RESULTS'] is the results section within each DETECTION element

    tls_detection_list = []
    for host in data:
        if 'DNS' not in host.keys():
            host_name = ''
        else:
            host_name = host['DNS']

        if type(host['DETECTION_LIST']['DETECTION']) is type([]):
            for detection in host['DETECTION_LIST']['DETECTION']:
                if detection['QID'] == '38116':
                    detection_data = process_38116_detection(detection)
                    detection_data['Hostname'] = host_name
                    detection_data['IP'] = host['IP']
                    tls_detection_list.append(detection_data)
                elif detection['QID'] == '38704':
                    detection_data = process_38704_detection(detection)
                    detection_data['Hostname'] = host_name
                    detection_data['IP'] = host['IP']
                    tls_detection_list.append(detection_data)
        else:
            if host['DETECTION_LIST']['DETECTION']['QID'] == '38116':
                detection_data = process_38116_detection(host['DETECTION_LIST']['DETECTION'])
                detection_data['Hostname'] = host_name
                detection_data['IP'] = host['IP']
                tls_detection_list.append(detection_data)
            elif host['DETECTION_LIST']['DETECTION']['QID'] == '38704':
                detection_data = process_38704_detection(host['DETECTION_LIST']['DETECTION'])
                detection_data['Hostname'] = host_name
                detection_data['IP'] = host['IP']
                tls_detection_list.append(detection_data)

    return tls_detection_list


def output_tls_inventory(wb: Workbook, inventory: list):
    output_38116_inventory(wb=wb, inventory=inventory)
    output_38704_inventory(wb=wb, inventory=inventory)


def output_38704_inventory(wb: Workbook, inventory: list):
    print('Building Key Exchange Methods worksheet')
    kex_ws = wb.create_sheet(title='Key Exchange Methods', index=1)
    kex_ws['A1'] = 'IP'
    kex_ws['B1'] = 'Hostname'
    kex_ws['C1'] = 'Port'
    kex_ws['D1'] = 'Last Detected'
    kex_ws['E1'] = 'Protocol'
    kex_ws['F1'] = 'Cipher'
    kex_ws['G1'] = 'Key Exchange Method'
    kex_ws['H1'] = 'Group'
    kex_ws['I1'] = 'Key Size'
    kex_ws['J1'] = 'Forward Secret'
    kex_ws['K1'] = 'Classical Strength'
    kex_ws['L1'] = 'Quantum Strength'

    row = 2

    for i in inventory:
        if (row-2) % 1000 == 0:
            print('.', end='')
        if 'TLS' not in i.keys():
            continue
        if '38704' not in i['TLS']:
            continue
        for protocol in i['TLS']['38704']['Protocols']:
            for kex_method in protocol['KeyExchangeMethods']:
                kex_ws.cell(row=row, column=1).value = i['IP']
                kex_ws.cell(row=row, column=2).value = i['Hostname']
                kex_ws.cell(row=row, column=3).value = int(i['TLS']['38704']['Port'])
                kex_ws.cell(row=row, column=4).value = i['TLS']['38704']['LastDetected']
                if 'Name' in protocol.keys():
                    kex_ws.cell(row=row, column=5).value = protocol['Name']
                if 'Cipher' in protocol.keys():
                    kex_ws.cell(row=row, column=6).value = kex_method['Cipher']
                kex_ws.cell(row=row, column=7).value = kex_method['Name']
                kex_ws.cell(row=row, column=8).value = kex_method['Group']
                if kex_method['KeySize'] == '':
                    kex_ws.cell(row=row, column=9).value = kex_method['KeySize']
                else:
                    # if type(kex_method['KeySize']) == str:
                    if isinstance(kex_method['KeySize'], str):
                        kex_ws.cell(row, column=9).value = kex_method['KeySize']
                    else:
                        kex_ws.cell(row=row, column=9).value = int(kex_method['KeySize'])
                kex_ws.cell(row=row, column=10).value = kex_method['ForwardSecret']
                kex_ws.cell(row=row, column=11).value = kex_method['ClassicalStrength']
                kex_ws.cell(row=row, column=12).value = kex_method['QuantumStrength']
                row += 1
    print('\nFinished')


def output_38116_inventory(wb: Workbook, inventory: list):
    print('Building TLS Cryptographic Inventory worksheet')
    tls_ws = wb.create_sheet(title='TLS Cryptographic Inventory', index=0)
    tls_ws['A1'] = 'IP'
    tls_ws['B1'] = 'Hostname'
    tls_ws['C1'] = 'Port'
    tls_ws['D1'] = 'Last Detected'
    tls_ws['E1'] = 'Protocol'
    tls_ws['F1'] = 'Compression Method'
    tls_ws['G1'] = 'Cipher'
    tls_ws['H1'] = 'Key Exchange'
    tls_ws['I1'] = 'Authentication'
    tls_ws['J1'] = 'MAC'
    tls_ws['K1'] = 'Encryption'
    tls_ws['L1'] = 'Key Strength'
    tls_ws['M1'] = 'Grade'

    row = 2
    for i in inventory:
        if (row-2) % 1000 == 0:
            print('.', end='')
        if 'TLS' not in i.keys():
            continue
        if '38116' not in i['TLS']:
            continue
        for protocol in i['TLS']['38116']['Protocols']:
            if protocol['Status'] == 'DISABLED':
                continue
            for cipher in protocol['Ciphersuite']:
                tls_ws.cell(row=row, column=1).value = i['IP']
                tls_ws.cell(row=row, column=2).value = i['Hostname']
                tls_ws.cell(row=row, column=3).value = int(i['TLS']['38116']['Port'])
                tls_ws.cell(row=row, column=4).value = i['TLS']['38116']['LastDetected']
                tls_ws.cell(row=row, column=5).value = protocol['Name']
                tls_ws.cell(row=row, column=6).value = protocol['CompressionMethod']
                tls_ws.cell(row=row, column=7).value = cipher['Name']
                tls_ws.cell(row=row, column=8).value = cipher['KeyExchange']
                tls_ws.cell(row=row, column=9).value = cipher['Authentication']
                tls_ws.cell(row=row, column=10).value = cipher['MAC']
                tls_ws.cell(row=row, column=11).value = cipher['Encryption']
                if cipher['KeyStrength'] == '':
                    tls_ws.cell(row=row, column=12).value = cipher['KeyStrength']
                else:
                    tls_ws.cell(row=row, column=12).value = int(cipher['KeyStrength'])
                tls_ws.cell(row=row, column=13).value = cipher['Grade']
                row += 1
    print('\nFinished')
