from openpyxl import Workbook


def process_ssh_detections(data: list):
    ssh_detection_list = []
    for host in data:
        if 'DNS' not in host.keys():
            host_name = ''
        else:
            host_name = host['DNS']

        if isinstance(host['DETECTION_LIST']['DETECTION'], list):
            for detection in host['DETECTION_LIST']['DETECTION']:
                if detection['QID'] == '38047':
                    detection_data = process_38047_detection(detection)
                    detection_data['Hostname'] = host_name
                    detection_data['IP'] = host['IP']
                    ssh_detection_list.append(detection_data)
        else:
            if host['DETECTION_LIST']['DETECTION']['QID'] == '38047':
                detection_data = process_38047_detection(host['DETECTION_LIST']['DETECTION'])
                detection_data['Hostname'] = host_name
                detection_data['IP'] = host['IP']
                ssh_detection_list.append(detection_data)
    return ssh_detection_list


def output_ssh_inventory(wb: Workbook, inventory: list):
    output_38047_inventory(wb=wb, inventory=inventory)


def process_38047_detection(detection: dict):
    detection_data = {
        'SSH': {
            '38047': {
                'Port': detection['PORT'],
                'LastDetected': detection['LAST_FOUND_DATETIME'],
                'SSH_Data': process_38047_results(detection['RESULTS'])
            }
        }
    }
    return detection_data


def process_38047_results(result_data: str):
    rows = result_data.split('\n')
    versions = {}
    version = {}
    for row in rows:
        cols = row.split('\t')
        if cols[0] == 'SSH1 supported':
            versions['1'] = {'Supported': cols[1]}
        elif cols[0] == 'SSH2 supported':
            versions['2'] = {'Supported': cols[1]}
        elif cols[0] == 'Supported authentification methods for SSH1' or \
                cols[0] == 'Supported authentication methods for SSH1':
            versions['1']['AuthMethods'] = cols[1]
        elif cols[0] == 'Supported ciphers for SSH1':
            versions['1']['Ciphers'] = cols[1]
        elif cols[0] == 'Supported key exchange algorithms for SSH2':
            versions['2']['KeyExchangeAlgorithms'] = cols[1]
        elif cols[0] == 'Supported decryption ciphers for SSH2':
            versions['2']['DecryptionCiphers'] = cols[1]
        elif cols[0] == 'Supported encryption ciphers for SSH2':
            versions['2']['EncryptionCiphers'] = cols[1]
        elif cols[0] == 'Supported decryption macs for SSH2':
            versions['2']['DecryptionMAC'] = cols[1]
        elif cols[0] == 'Supported encryption macs for SSH2':
            versions['2']['EncryptionMAC'] = cols[1]
        elif cols[0] == 'Supported authentification methods for SSH2' or \
                cols[0] == 'Supported authentication methods for SSH2':
            versions['2']['AuthMethods'] = cols[1]
        elif cols[0] == 'Supported host key algorithms for SSH2':
            versions['2']['HostKeyAlgorithms'] = cols[1]
        elif cols[0] == 'Supported decompression for SSH2':
            versions['2']['Decompression'] = cols[1]
        elif cols[0] == 'Supported compression for SSH2':
            versions['2']['Compression'] = cols[1]

    return versions


def output_38047_inventory(wb: Workbook, inventory: list):
    print('Building SSH Information worksheet', index=3)
    ssh_ws = wb.create_sheet(title='SSH Information')
    ssh_ws['A1'] = 'IP'
    ssh_ws['B1'] = 'Hostname'
    ssh_ws['C1'] = 'Port'
    ssh_ws['D1'] = 'Last Detected'
    ssh_ws['E1'] = 'SSH1 Supported'
    ssh_ws['F1'] = 'SSH1 Authentication Methods'
    ssh_ws['G1'] = 'SSH1 Ciphers'
    ssh_ws['H1'] = 'SSH2 Supported'
    ssh_ws['I1'] = 'SSH2 Key Exchange Algorithms'
    ssh_ws['J1'] = 'SSH2 Decryption Ciphers'
    ssh_ws['K1'] = 'SSH2 Encryption Ciphers'
    ssh_ws['L1'] = 'SSH2 Decryption MAC'
    ssh_ws['M1'] = 'SSH2 Encryption MAC'
    ssh_ws['N1'] = 'SSH2 Authentication Methods'
    ssh_ws['O1'] = 'SSH2 Host Key Algorithms'
    ssh_ws['P1'] = 'SSH2 Decompression'
    ssh_ws['Q1'] = 'SSH2 Compression'

    row = 2
    for i in inventory:
        if (row-2) % 1000 == 0:
            print('.', end='')
        if 'SSH' not in i.keys():
            continue
        if '38047' not in i['SSH']:
            continue
        ssh_ws.cell(row=row, column=1).value = i['IP']
        ssh_ws.cell(row=row, column=2).value = i['Hostname']
        ssh_ws.cell(row=row, column=3).value = i['SSH']['38047']['Port']
        ssh_ws.cell(row=row, column=4).value = i['SSH']['38047']['LastDetected']
        ssh_ws.cell(row=row, column=5).value = i['SSH']['38047']['SSH_Data']['1']['Supported']
        if i['SSH']['38047']['SSH_Data']['1']['Supported'] == 'no':
            ssh_ws.cell(row=row, column=6).value = '-'
            ssh_ws.cell(row=row, column=7).value = '-'
        else:
            ssh_ws.cell(row=row, column=6).value = i['SSH']['38047']['SSH_Data']['1']['AuthMethods']
            ssh_ws.cell(row=row, column=7).value = i['SSH']['38047']['SSH_Data']['1']['Ciphers']
        ssh_ws.cell(row=row, column=8).value = i['SSH']['38047']['SSH_Data']['2']['Supported']
        if i['SSH']['38047']['SSH_Data']['2']['Supported'] == 'no':
            col = 9
            while col < 18:
                ssh_ws.cell(row=row, column=col).value = '-'
                col += 1
        else:
            ssh_ws.cell(row=row, column=9).value = i['SSH']['38047']['SSH_Data']['2']['KeyExchangeAlgorithms']
            ssh_ws.cell(row=row, column=10).value = i['SSH']['38047']['SSH_Data']['2']['DecryptionCiphers']
            ssh_ws.cell(row=row, column=11).value = i['SSH']['38047']['SSH_Data']['2']['EncryptionCiphers']
            ssh_ws.cell(row=row, column=12).value = i['SSH']['38047']['SSH_Data']['2']['DecryptionMAC']
            ssh_ws.cell(row=row, column=13).value = i['SSH']['38047']['SSH_Data']['2']['EncryptionMAC']
            ssh_ws.cell(row=row, column=14).value = i['SSH']['38047']['SSH_Data']['2']['AuthMethods']
            ssh_ws.cell(row=row, column=15).value = i['SSH']['38047']['SSH_Data']['2']['HostKeyAlgorithms']
            ssh_ws.cell(row=row, column=16).value = i['SSH']['38047']['SSH_Data']['2']['Decompression']
            ssh_ws.cell(row=row, column=17).value = i['SSH']['38047']['SSH_Data']['2']['Compression']

        row += 1
    print('\nFinished')
